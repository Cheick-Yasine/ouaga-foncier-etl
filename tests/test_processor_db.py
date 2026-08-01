"""Tests de la base maître PostgreSQL (upsert, export Excel, détection de dérive).

Tous les tests de ce fichier qui touchent réellement une base utilisent la
fixture `base_test_isolee` (voir conftest.py) et sont automatiquement skippés
si TEST_DATABASE_URL n'est pas définie ou que le serveur est injoignable -
voir README.md, section Tests, pour la configuration locale.
"""

from __future__ import annotations

import openpyxl
import psycopg

import processor

ANNONCE_1 = {
    "id": "p1", "groupe_nom": "Test", "url": "https://x/1", "date_publication": None,
    "date_incertaine": True, "type_bien": "parcelle", "quartier_zone": "Ouaga 2000",
    "superficie_m2": 600, "prix_fcfa": 15000000, "statut_document": "Titre Foncier",
    "contacts_whatsapp": ["70123456"], "mots_cles_pertinents": ["parcelle"],
    "resume_court": "Résumé 1", "texte_nettoye": "texte 1",
}
ANNONCE_2 = {
    "id": "p2", "groupe_nom": "Test", "url": "https://x/2", "date_publication": None,
    "date_incertaine": True, "type_bien": "villa", "quartier_zone": "Pissy",
    "superficie_m2": 300, "prix_fcfa": 25000000, "statut_document": "PUH",
    "contacts_whatsapp": ["76000000"], "mots_cles_pertinents": ["villa"],
    "resume_court": "Résumé 2", "texte_nettoye": "texte 2",
}


class TestUpsertAnnonces:
    def test_insere_de_nouvelles_annonces(self, base_test_isolee):
        n = processor.upsert_annonces([ANNONCE_1, ANNONCE_2], dsn=base_test_isolee)
        assert n == 2

        with psycopg.connect(base_test_isolee) as conn:
            lignes = conn.execute("SELECT id, prix_fcfa FROM annonces ORDER BY id").fetchall()
        assert lignes == [("p1", 15000000), ("p2", 25000000)]

    def test_reinsertion_met_a_jour_sans_dupliquer(self, base_test_isolee):
        processor.upsert_annonces([ANNONCE_1], dsn=base_test_isolee)
        annonce_maj = {**ANNONCE_1, "prix_fcfa": 12000000}
        processor.upsert_annonces([annonce_maj], dsn=base_test_isolee)

        with psycopg.connect(base_test_isolee) as conn:
            lignes = conn.execute("SELECT id, prix_fcfa FROM annonces").fetchall()
        assert lignes == [("p1", 12000000)]  # une seule ligne, prix mis à jour

    def test_premiere_collecte_nest_pas_ecrasee_par_une_maj(self, base_test_isolee):
        processor.upsert_annonces([ANNONCE_1], dsn=base_test_isolee)
        with psycopg.connect(base_test_isolee) as conn:
            premiere_avant = conn.execute(
                "SELECT premiere_collecte FROM annonces WHERE id='p1'"
            ).fetchone()[0]

        processor.upsert_annonces([{**ANNONCE_1, "prix_fcfa": 1}], dsn=base_test_isolee)
        with psycopg.connect(base_test_isolee) as conn:
            premiere_apres = conn.execute(
                "SELECT premiere_collecte FROM annonces WHERE id='p1'"
            ).fetchone()[0]
        assert premiere_avant == premiere_apres

    def test_liste_vide_ne_cree_rien(self, base_test_isolee):
        n = processor.upsert_annonces([], dsn=base_test_isolee)
        assert n == 0


class TestExporterXlsxDepuisDb:
    def test_genere_un_fichier_avec_les_bonnes_lignes(self, base_test_isolee, tmp_path):
        xlsx = tmp_path / "annonces.xlsx"
        processor.upsert_annonces([ANNONCE_1, ANNONCE_2], dsn=base_test_isolee)

        chemin = processor.exporter_xlsx_depuis_db(dsn=base_test_isolee, chemin_xlsx=xlsx)

        assert chemin == xlsx
        classeur = openpyxl.load_workbook(xlsx)
        feuille = classeur.active
        assert feuille.max_row == 3  # en-tête + 2 lignes
        assert feuille.cell(row=1, column=1).value == "id"

    def test_reexport_ecrase_le_meme_fichier(self, base_test_isolee, tmp_path):
        xlsx = tmp_path / "annonces.xlsx"
        processor.upsert_annonces([ANNONCE_1], dsn=base_test_isolee)
        processor.exporter_xlsx_depuis_db(dsn=base_test_isolee, chemin_xlsx=xlsx)
        processor.upsert_annonces([ANNONCE_2], dsn=base_test_isolee)
        processor.exporter_xlsx_depuis_db(dsn=base_test_isolee, chemin_xlsx=xlsx)

        classeur = openpyxl.load_workbook(xlsx)
        assert classeur.active.max_row == 3  # en-tête + 2 lignes, un seul fichier


class TestDetecterDerive:
    def test_pas_dassez_dhistorique_ne_declenche_rien(self, base_test_isolee):
        processor.enregistrer_run("daily", 100, 20, 10, dsn=base_test_isolee)
        processor.enregistrer_run("daily", 100, 20, 10, dsn=base_test_isolee)
        # seulement 2 runs enregistrés, minimum_historique par défaut = 3
        assert processor.detecter_derive(1, mode="daily", dsn=base_test_isolee) is None

    def test_volume_normal_ne_declenche_rien(self, base_test_isolee):
        for _ in range(5):
            processor.enregistrer_run("daily", 100, 20, 10, dsn=base_test_isolee)
        assert processor.detecter_derive(9, mode="daily", dsn=base_test_isolee) is None

    def test_volume_anormalement_bas_declenche_une_alerte(self, base_test_isolee):
        for _ in range(5):
            processor.enregistrer_run("daily", 100, 20, 10, dsn=base_test_isolee)
        alerte = processor.detecter_derive(0, mode="daily", dsn=base_test_isolee)
        assert alerte is not None
        assert "anormalement bas" in alerte

    def test_mode_backfill_nest_jamais_evalue(self, base_test_isolee):
        for _ in range(5):
            processor.enregistrer_run("daily", 100, 20, 10, dsn=base_test_isolee)
        assert processor.detecter_derive(0, mode="backfill", dsn=base_test_isolee) is None

    def test_base_injoignable_ne_declenche_rien(self):
        # DSN pointant vers un port fermé en local : échec de connexion rapide
        # et volontaire (ECONNREFUSED), sans dépendre d'un serveur de test réel.
        dsn_invalide = "postgresql://user:pass@localhost:1/inexistante"
        assert processor.detecter_derive(0, mode="daily", dsn=dsn_invalide) is None


class TestExecuterTraitementResultat:
    async def test_retourne_un_resultat_traitement_complet(
        self, monkeypatch, repertoires_isoles, base_test_isolee, tmp_path
    ):
        fichier = tmp_path / "posts.json"
        fichier.write_text(
            '[{"id": "p1", "groupe_nom": "T", "url": "https://x", '
            '"texte": "Terrain à vendre Ouaga 2000, 600m2, titre foncier, prix 15000000"}]',
            encoding="utf-8",
        )

        async def _fausse_structuration(_candidats, api_key=None):
            return [{**ANNONCE_1}], []

        monkeypatch.setattr(processor, "structurer_lot", _fausse_structuration)

        resultat = await processor.executer_traitement([fichier], mode="daily")

        assert isinstance(resultat, processor.ResultatTraitement)
        assert resultat.nb_posts_bruts == 1
        assert resultat.nb_valides == 1
        assert resultat.database_url == base_test_isolee
        assert resultat.chemin_xlsx.exists()
        assert resultat.chemin_csv_run.exists()
